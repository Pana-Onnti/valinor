"""
Upload router — File upload with validation and storage.

Accepts CSV and Excel files for analysis, validates extension/size/content,
detects Excel sheet names, saves to tenant-isolated storage, and registers
upload metadata in PostgreSQL (uploaded_files table with RLS).

Adds a /process endpoint that converts an uploaded file to SQLite for
downstream analysis (VAL-84), and preview/schema endpoints (VAL-85).

Refs: VAL-83, VAL-84, VAL-85
"""

import io
import os
import sqlite3
import uuid
from pathlib import Path
from typing import Optional

import pandas as pd
import structlog
from fastapi import APIRouter, File, HTTPException, Query, Request, UploadFile
from sqlalchemy import create_engine, text

from api.models import (
    ColumnInfo, PreviewResponse, ProcessResponse, SchemaResponse, SchemaTable,
    TableInfo, UploadResponse,
)

logger = structlog.get_logger()

router = APIRouter(prefix="/api/upload", tags=["Upload"])

# ── Configuration ─────────────────────────────────────────────────────────────

ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
MAX_UPLOAD_SIZE = int(os.getenv("MAX_UPLOAD_SIZE_MB", "50")) * 1024 * 1024

_DEFAULT_UPLOAD_DIR = "/tmp/valinor/uploads"
UPLOAD_DIR = Path(os.getenv("UPLOAD_DIR", _DEFAULT_UPLOAD_DIR))

_DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://valinor:valinor_secret@localhost:5450/valinor_metadata")
_metadata_engine = create_engine(_DATABASE_URL, pool_pre_ping=True, pool_size=3)


# ── Upload DB helpers ─────────────────────────────────────────────────────────

def _db_insert_upload(upload_id: str, tenant_id: str, client_name: str,
                      filename: str, stored_path: str, file_size: int,
                      file_type: str) -> None:
    with _metadata_engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO uploaded_files
                (id, tenant_id, client_name, original_filename, stored_path, file_size, file_type)
            VALUES
                (:id, :tenant_id::uuid, :client_name, :filename, :stored_path, :file_size, :file_type)
        """), {
            "id": upload_id,
            "tenant_id": tenant_id,
            "client_name": client_name,
            "filename": filename,
            "stored_path": stored_path,
            "file_size": file_size,
            "file_type": file_type,
        })


def _db_get_upload(upload_id: str) -> Optional[dict]:
    with _metadata_engine.connect() as conn:
        row = conn.execute(text("""
            SELECT id, tenant_id, client_name, original_filename, stored_path,
                   file_size, file_type, status, uploaded_at, processed_at
            FROM uploaded_files WHERE id = :id
        """), {"id": upload_id}).mappings().first()
        if row is None:
            return None
        return dict(row)


def _db_mark_processed(upload_id: str, db_path: str) -> None:
    with _metadata_engine.begin() as conn:
        conn.execute(text("""
            UPDATE uploaded_files
            SET status = 'processed', processed_at = now()
            WHERE id = :id
        """), {"id": upload_id})


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_extension(filename: str) -> str:
    """Return the lowercased file extension including the leading dot."""
    return Path(filename).suffix.lower()


def _get_file_type(extension: str) -> str:
    """Map extension to a canonical file_type string."""
    return extension.lstrip(".")


def _detect_excel_sheets(content: bytes) -> list[str]:
    """
    Load an Excel workbook in read-only mode and return its sheet names.

    Returns an empty list if the file cannot be parsed as a valid workbook.
    """
    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        return list(sheets)
    except Exception as exc:
        logger.warning("upload.excel_sheet_detection_failed", error=str(exc))
        return []


def _validate_csv_content(content: bytes) -> None:
    """
    Verify that the CSV bytes look like text and contain at least one data row.

    Raises HTTPException 400 if the check fails.
    """
    try:
        text = content.decode("utf-8", errors="replace")
    except Exception:
        raise HTTPException(status_code=400, detail="CSV file is not valid UTF-8 text")

    lines = [ln for ln in text.splitlines() if ln.strip()]
    if len(lines) < 2:
        raise HTTPException(
            status_code=400,
            detail="CSV file must contain a header row and at least one data row",
        )


def _tenant_id_from_request(request: Request) -> str:
    """
    Return the tenant_id set by TenantMiddleware, falling back to the
    default development tenant if not present.
    """
    return getattr(request.state, "tenant_id", None) or "00000000-0000-0000-0000-000000000001"


def _build_storage_path(tenant_id: str, client_name: str, upload_id: str, filename: str) -> Path:
    """Construct and create the tenant-isolated storage path."""
    dest_dir = UPLOAD_DIR / tenant_id / client_name
    dest_dir.mkdir(parents=True, exist_ok=True)
    safe_filename = Path(filename).name  # strip any path traversal
    return dest_dir / f"{upload_id}_{safe_filename}"


# ── Endpoint ──────────────────────────────────────────────────────────────────

@router.post("/{client_name}", response_model=UploadResponse)
async def upload_file(
    client_name: str,
    request: Request,
    file: UploadFile = File(...),
) -> UploadResponse:
    """
    Upload a CSV or Excel file for analysis.

    Validations (in order):
    1. Extension must be .csv, .xlsx, or .xls  → 400
    2. File must not be empty                   → 400
    3. Size must not exceed MAX_UPLOAD_SIZE      → 413
    4. Content-specific checks (CSV rows / Excel sheets)

    The file is saved under UPLOAD_DIR/{tenant_id}/{client_name}/{upload_id}_{filename}.
    Metadata is recorded in the in-memory registry until the DB migration lands.
    """
    filename = file.filename or "upload"

    # 1. Validate extension
    extension = _get_extension(filename)
    if extension not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Unsupported file type '{extension}'. "
                f"Allowed types: {', '.join(sorted(ALLOWED_EXTENSIONS))}"
            ),
        )

    # 2. Read content (async — does not block event loop)
    content = await file.read()

    # 3. Empty file check
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    # 4. Size check
    if len(content) > MAX_UPLOAD_SIZE:
        max_mb = MAX_UPLOAD_SIZE // (1024 * 1024)
        raise HTTPException(
            status_code=413,
            detail=f"File exceeds the maximum allowed size of {max_mb} MB",
        )

    # 5. Content-specific validation and sheet detection
    sheets: list[str] = []
    if extension == ".csv":
        _validate_csv_content(content)
    else:
        # .xlsx / .xls
        sheets = _detect_excel_sheets(content)
        if not sheets:
            raise HTTPException(
                status_code=400,
                detail="Could not read any sheets from the Excel file. Ensure the file is a valid workbook.",
            )

    # 6. Tenant isolation
    tenant_id = _tenant_id_from_request(request)

    # 7. Persist to storage
    upload_id = str(uuid.uuid4())
    storage_path = _build_storage_path(tenant_id, client_name, upload_id, filename)

    try:
        storage_path.write_bytes(content)
    except OSError as exc:
        logger.error(
            "upload.storage_write_failed",
            path=str(storage_path),
            error=str(exc),
        )
        raise HTTPException(status_code=500, detail="Failed to save uploaded file")

    # 8. Record metadata in PostgreSQL
    file_type = _get_file_type(extension)
    _db_insert_upload(
        upload_id=upload_id,
        tenant_id=tenant_id,
        client_name=client_name,
        filename=filename,
        stored_path=str(storage_path),
        file_size=len(content),
        file_type=file_type,
    )

    logger.info(
        "upload.received",
        upload_id=upload_id,
        tenant_id=tenant_id,
        client_name=client_name,
        filename=filename,
        size_bytes=len(content),
        file_type=file_type,
        sheets=sheets,
    )

    return UploadResponse(
        upload_id=upload_id,
        filename=filename,
        size_bytes=len(content),
        file_type=file_type,
        sheets=sheets,
        status="pending",
    )


# ── Process endpoint (VAL-84) ──────────────────────────────────────────────────

@router.post("/{upload_id}/process", response_model=ProcessResponse)
async def process_upload(upload_id: str) -> ProcessResponse:
    """
    Convert a previously uploaded CSV/Excel file to a SQLite database.
    """
    meta = _db_get_upload(upload_id)
    if meta is None:
        raise HTTPException(status_code=404, detail=f"Upload '{upload_id}' not found")

    if meta.get("status") == "processed":
        raise HTTPException(
            status_code=400,
            detail=f"Upload '{upload_id}' has already been processed",
        )

    from api.services.file_ingestion import FileIngestionService

    service = FileIngestionService()

    try:
        result = service.convert_to_sqlite(
            file_path=meta["stored_path"],
            file_type=meta["file_type"],
            client_name=meta["client_name"],
            tenant_id=str(meta["tenant_id"]),
            upload_id=upload_id,
        )
    except FileNotFoundError as exc:
        logger.error("upload.process.file_not_found", upload_id=upload_id, error=str(exc))
        raise HTTPException(status_code=500, detail=f"Source file missing: {exc}")
    except Exception as exc:
        logger.error("upload.process.failed", upload_id=upload_id, error=str(exc))
        raise HTTPException(
            status_code=500,
            detail=f"File conversion failed: {exc}",
        )

    _db_mark_processed(upload_id, result["db_path"])

    logger.info(
        "upload.process.done",
        upload_id=upload_id,
        db_path=result["db_path"],
        tables=[t["name"] for t in result["tables"]],
    )

    return ProcessResponse(
        upload_id=upload_id,
        status="processed",
        db_path=result["db_path"],
        tables=[
            TableInfo(
                name=t["name"],
                row_count=t["row_count"],
                columns=t["columns"],
            )
            for t in result["tables"]
        ],
    )


# ── Preview endpoint (VAL-85) ─────────────────────────────────────────────────

def _read_csv_with_encoding(path: Path, nrows: int) -> pd.DataFrame:
    """Try UTF-8 first, fall back to latin-1 on encoding errors."""
    try:
        return pd.read_csv(path, nrows=nrows, encoding="utf-8")
    except UnicodeDecodeError:
        return pd.read_csv(path, nrows=nrows, encoding="latin-1")


def _count_csv_data_rows(path: Path) -> int:
    """Count data rows in a CSV (header not included) without loading all data."""
    try:
        for enc in ("utf-8", "latin-1"):
            try:
                with open(path, encoding=enc) as fh:
                    count = sum(1 for line in fh if line.strip()) - 1
                return max(count, 0)
            except UnicodeDecodeError:
                continue
        return 0
    except OSError:
        return 0


def _count_excel_rows(path: Path, sheet: str | int) -> int:
    """Count data rows in an Excel sheet using openpyxl read-only mode."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if isinstance(sheet, int):
            ws = wb.worksheets[sheet]
        else:
            ws = wb[sheet]
        row_count = max((ws.max_row or 1) - 1, 0)
        wb.close()
        return row_count
    except Exception as exc:
        logger.warning("upload.excel_row_count_failed", error=str(exc))
        return 0


def _build_column_info(df: pd.DataFrame) -> list[ColumnInfo]:
    """Build ColumnInfo list from a DataFrame."""
    columns: list[ColumnInfo] = []
    for col in df.columns:
        series = df[col]
        null_count = int(series.isna().sum())
        non_null = series.dropna()
        sample = str(non_null.iloc[0]) if len(non_null) > 0 else None
        columns.append(
            ColumnInfo(
                name=str(col),
                dtype=str(series.dtype),
                nulls=null_count,
                sample=sample,
            )
        )
    return columns


@router.get("/{upload_id}/preview", response_model=PreviewResponse)
async def preview_upload(
    upload_id: str,
    rows: int = Query(default=20, ge=1, le=100, description="Number of rows to preview (max 100)"),
    sheet: Optional[str] = Query(default=None, description="Sheet name (Excel only)"),
) -> PreviewResponse:
    """
    Preview the first N rows of an uploaded file.
    """
    meta = _db_get_upload(upload_id)
    if meta is None:
        raise HTTPException(status_code=404, detail=f"Upload '{upload_id}' not found")

    path = Path(meta["stored_path"])
    if not path.exists():
        raise HTTPException(status_code=404, detail="Stored file not found on disk")

    file_type = meta["file_type"]
    filename = meta["original_filename"]
    # Detect sheets on-the-fly for Excel files (not stored in DB)
    sheets_available: list[str] = _detect_excel_sheets(path.read_bytes()) if file_type in ("xlsx", "xls") else []
    active_sheet: Optional[str] = None

    if file_type == "csv":
        df = _read_csv_with_encoding(path, nrows=rows)
        total_rows = _count_csv_data_rows(path)
    else:
        if sheet is not None:
            if sheet not in sheets_available:
                raise HTTPException(
                    status_code=400,
                    detail=f"Sheet '{sheet}' not found. Available sheets: {sheets_available}",
                )
            active_sheet = sheet
            sheet_arg: str | int = sheet
        else:
            active_sheet = sheets_available[0] if sheets_available else None
            sheet_arg = 0

        df = pd.read_excel(path, sheet_name=sheet_arg, nrows=rows)
        total_rows = _count_excel_rows(path, sheet_arg)

    column_info = _build_column_info(df)

    rows_data = df.where(df.notna(), other=None).to_dict(orient="records")
    serialized_rows = [
        {k: (str(v) if v is not None else None) for k, v in row.items()}
        for row in rows_data
    ]

    logger.info(
        "upload.preview",
        upload_id=upload_id,
        rows_requested=rows,
        rows_returned=len(serialized_rows),
        sheet=active_sheet,
    )

    return PreviewResponse(
        upload_id=upload_id,
        filename=filename,
        sheet=active_sheet,
        sheets_available=sheets_available,
        total_rows=total_rows,
        columns=column_info,
        rows=serialized_rows,
    )


# ── Schema endpoint (VAL-85) ──────────────────────────────────────────────────

@router.get("/{upload_id}/schema", response_model=SchemaResponse)
async def get_upload_schema(upload_id: str) -> SchemaResponse:
    """
    Get schema information from a processed upload's SQLite database.
    """
    meta = _db_get_upload(upload_id)
    if meta is None:
        raise HTTPException(status_code=404, detail=f"Upload '{upload_id}' not found")

    upload_status = meta.get("status", "pending")
    if upload_status != "processed":
        raise HTTPException(
            status_code=400,
            detail=f"Upload is not yet processed (current status: '{upload_status}'). "
                   "Schema is only available after processing completes.",
        )

    # Derive SQLite path from storage convention: {tenant}/{client}/processed/{id}.db
    stored = Path(meta["stored_path"])
    sqlite_path = stored.parent / "processed" / f"{upload_id}.db"
    if not sqlite_path or not sqlite_path.exists():
        raise HTTPException(
            status_code=404,
            detail="Processed SQLite file not found. Re-process the upload.",
        )

    tables: list[SchemaTable] = []
    try:
        con = sqlite3.connect(str(sqlite_path))
        cur = con.cursor()

        cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        table_names = [row[0] for row in cur.fetchall()]

        for table_name in table_names:
            cur.execute(f"PRAGMA table_info([{table_name}])")
            pragma_rows = cur.fetchall()
            col_info = [
                ColumnInfo(name=row[1], dtype=row[2] or "TEXT")
                for row in pragma_rows
            ]

            cur.execute(f"SELECT COUNT(*) FROM [{table_name}]")
            row_count = cur.fetchone()[0]

            tables.append(SchemaTable(name=table_name, row_count=row_count, columns=col_info))

        con.close()
    except sqlite3.Error as exc:
        logger.error("upload.schema_sqlite_error", upload_id=upload_id, error=str(exc))
        raise HTTPException(status_code=500, detail=f"Failed to read SQLite schema: {exc}")

    logger.info("upload.schema", upload_id=upload_id, tables=[t.name for t in tables])

    return SchemaResponse(upload_id=upload_id, tables=tables)
